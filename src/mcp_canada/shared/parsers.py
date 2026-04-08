"""Shared file parser for XLSX, XLS, and CSV files from government URLs.

Public API:
    fetch_and_parse(url, sheet, skip_rows, ttl) -> (list[dict], was_cached)
"""

from __future__ import annotations

import csv
import re
from io import BytesIO, StringIO
from typing import Any

import httpx

from mcp_canada.shared.cache import cached_fetch

# Regex to collapse non-alphanumeric runs to underscores
_NON_ALNUM = re.compile(r"[^a-z0-9]+")


def _normalize_key(header: str) -> str:
    """Normalize a column header to a snake_case identifier.

    Examples:
        "Country of Citizenship" -> "country_of_citizenship"
        "  Year " -> "year"
        "123col" -> "col_123col"
        "" -> "col"
        "a/b-c d" -> "a_b_c_d"
    """
    key = header.strip().lower()
    key = _NON_ALNUM.sub("_", key)
    key = key.strip("_")
    if not key:
        return "col"
    if key[0].isdigit():
        key = f"col_{key}"
    return key


def _mask_privacy(value: Any) -> Any:
    """Convert privacy-masked '--' values to None.

    Returns None for '--' strings (with optional surrounding whitespace),
    and None for pandas NaN/NaT/NA if pandas is available. All other
    values are returned unchanged.
    """
    if value is None:
        return None
    if isinstance(value, str):
        if value.strip() == "--":
            return None
        return value
    # Handle pandas NA types if pandas is available
    try:
        import pandas as pd  # noqa: PLC0415

        if pd.isna(value):
            return None
    except (ImportError, TypeError, ValueError):
        pass
    return value


def _parse_xlsx_pandas(
    content: bytes,
    sheet: str | int = 0,
    skip_rows: int = 0,
) -> list[dict[str, Any]]:
    """Parse XLSX bytes using pandas.

    Args:
        content: Raw XLSX file bytes.
        sheet: Sheet name or 0-based index.
        skip_rows: Number of rows to skip before the header row.

    Returns:
        list of dicts with normalized snake_case keys and privacy-masked values.
    """
    import pandas as pd  # noqa: PLC0415

    df = pd.read_excel(BytesIO(content), sheet_name=sheet, skiprows=skip_rows)
    # Rename columns to normalized keys
    df.columns = [_normalize_key(str(col)) for col in df.columns]
    records = df.to_dict(orient="records")
    return [{k: _mask_privacy(v) for k, v in row.items()} for row in records]


def _parse_xlsx_openpyxl(
    content: bytes,
    sheet: str | int = 0,
    skip_rows: int = 0,
) -> list[dict[str, Any]]:
    """Parse XLSX bytes using openpyxl (fallback when pandas is unavailable).

    Args:
        content: Raw XLSX file bytes.
        sheet: Sheet name or 0-based index.
        skip_rows: Number of rows to skip before the header row.

    Returns:
        list of dicts with normalized snake_case keys and privacy-masked values.
    """
    import openpyxl  # noqa: PLC0415

    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        if isinstance(sheet, int):
            ws = wb.worksheets[sheet]
        else:
            ws = wb[sheet]

        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        return []

    # Apply skip_rows before header
    rows = rows[skip_rows:]
    if not rows:
        return []

    headers = [_normalize_key(str(h) if h is not None else "") for h in rows[0]]
    result: list[dict[str, Any]] = []
    for row in rows[1:]:
        record = {headers[i]: _mask_privacy(v) for i, v in enumerate(row) if i < len(headers)}
        result.append(record)
    return result


def _parse_xlsx(
    content: bytes,
    sheet: str | int = 0,
    skip_rows: int = 0,
) -> list[dict[str, Any]]:
    """Parse XLSX bytes using pandas (preferred) or openpyxl fallback.

    Attempts pandas first for better multi-sheet, encoding, and type handling.
    Falls back to openpyxl + stdlib if pandas is not installed.
    """
    try:
        import pandas  # noqa: F401, PLC0415

        return _parse_xlsx_pandas(content, sheet, skip_rows)
    except ImportError:
        return _parse_xlsx_openpyxl(content, sheet, skip_rows)


def _parse_xls(
    content: bytes,
    sheet: str | int = 0,
    skip_rows: int = 0,
) -> list[dict[str, Any]]:
    """Parse legacy XLS bytes using xlrd.

    Raises:
        ImportError: If xlrd is not installed.
    """
    try:
        import xlrd  # noqa: PLC0415
    except ImportError as exc:
        raise ImportError(
            "Install xlrd for .xls support: pip install mcp-canada[ircc]"
        ) from exc

    wb = xlrd.open_workbook(file_contents=content)
    if isinstance(sheet, int):
        ws = wb.sheet_by_index(sheet)
    else:
        ws = wb.sheet_by_name(sheet)

    if ws.nrows <= skip_rows:
        return []

    headers = [_normalize_key(str(ws.cell_value(skip_rows, c))) for c in range(ws.ncols)]
    result: list[dict[str, Any]] = []
    for r in range(skip_rows + 1, ws.nrows):
        record = {headers[c]: _mask_privacy(ws.cell_value(r, c)) for c in range(ws.ncols)}
        result.append(record)
    return result


def _parse_ircc_xlsx(
    content: bytes,
    skip_rows: int,
    header_rows: int,
    label_cols: int,
    sheet: str | int = 0,
) -> list[dict[str, Any]]:
    """Parse IRCC XLSX bytes with multi-row merged headers into a flat list of dicts.

    IRCC XLSX files use 3-5 header rows encoding a Year > Quarter > Month hierarchy
    via merged cells. This function uses openpyxl to forward-fill merged cells and
    build flat composite column names like "col_2015_q1_jan" or "col_2015_total".

    Args:
        content: Raw XLSX file bytes.
        skip_rows: Number of rows to skip before the header block (title, blank rows).
        header_rows: Number of consecutive header rows to combine into column names.
        label_cols: Number of label columns at the left (e.g. "Country", "Province").
        sheet: Sheet name or 0-based index.

    Returns:
        list of dicts with normalized snake_case keys and privacy-masked values.
    """
    import openpyxl  # noqa: PLC0415

    wb = openpyxl.load_workbook(BytesIO(content), read_only=False, data_only=True)
    try:
        if isinstance(sheet, int):
            ws = wb.worksheets[sheet]
        else:
            ws = wb[sheet]

        rows = [
            tuple(cell.value for cell in row)
            for row in ws.iter_rows()
        ]
    finally:
        wb.close()

    if not rows:
        return []

    # Skip title/blank rows
    rows = rows[skip_rows:]
    if len(rows) < header_rows + 1:
        return []

    # Extract the header rows block
    header_block_raw = [list(row) for row in rows[:header_rows]]
    data_rows = rows[header_rows:]

    # Forward-fill None values in each header row (left-to-right) into a COPY.
    # This is used to propagate year/quarter context to their month sub-columns.
    # We keep the original (raw) values to know which columns had explicit content.
    header_block_filled = [list(hrow) for hrow in header_block_raw]
    for hrow in header_block_filled:
        last_val: Any = None
        for i, val in enumerate(hrow):
            if val is not None:
                last_val = val
            elif last_val is not None:
                hrow[i] = last_val

    # Build composite column names.
    # Strategy for temporal columns: use the filled value from each row,
    # BUT only if the column EITHER had an explicit value in that row OR
    # the row above (after fill) had a value for this column — unless the
    # immediately-higher row already had an EXPLICIT value (meaning this
    # column is "owned" by that level, not by forward-fill overflow).
    #
    # Simpler approach: for each temporal column, collect parts from each
    # header row using the FILLED value, but deduplicate consecutive equal
    # values AND skip values that are identical to the previous row's filled
    # value when the current row's RAW value was None.
    n_cols = len(header_block_filled[0]) if header_block_filled else 0
    headers: list[str] = []
    for col_idx in range(n_cols):
        if col_idx < label_cols:
            # Label columns: use the first filled row's value.
            # When multiple label columns share a merged header cell, the first
            # col gets the base name; subsequent cols get a numeric suffix to
            # ensure uniqueness (e.g. "gender_and_province_1", "gender_and_province_2").
            label_val = header_block_filled[0][col_idx]
            if label_val is not None:
                base_name = _normalize_key(str(label_val).strip())
            else:
                base_name = "label"
            if col_idx == 0:
                headers.append(base_name)
            else:
                # Check if this name already exists (from merged header)
                if base_name in headers:
                    headers.append(f"{base_name}_{col_idx + 1}")
                else:
                    headers.append(base_name)
        else:
            # Temporal columns: join values from each header row into composite names.
            #
            # Rule: use the forward-filled value for ALL rows EXCEPT the last header
            # row (the most granular level, e.g. month). For the last row, only include
            # its value if the RAW (pre-fill) value was explicitly set.
            #
            # This correctly handles "Year Total" columns where:
            #   - row N-1 has "Year Total" explicitly (quarter level)
            #   - row N (month level) is blank — so no month suffix is added
            # And handles "Feb" columns where:
            #   - row N-1 forward-fills "Q1" from the merged quarter cell
            #   - row N has "Feb" explicitly — so Q1 + Feb are both included
            last_row_idx = len(header_block_raw) - 1
            parts = []
            for row_idx, (raw_hrow, filled_hrow) in enumerate(
                zip(header_block_raw, header_block_filled)
            ):
                filled_val = filled_hrow[col_idx]
                if filled_val is None:
                    continue
                part = str(filled_val).strip()
                if not part:
                    continue

                if row_idx < last_row_idx:
                    # Upper rows (year, quarter): always include forward-fill context
                    if not parts or parts[-1] != part:
                        parts.append(part)
                else:
                    # Last row (month): only include if explicitly set in raw data
                    raw_val = raw_hrow[col_idx]
                    if raw_val is not None:
                        if not parts or parts[-1] != part:
                            parts.append(part)

            if parts:
                # Deduplicate adjacent parts (e.g. ["2015", "2015 Total"] -> ["2015", "Total"])
                deduped: list[str] = []
                for p in parts:
                    # If part starts with previous part text, strip the prefix
                    if deduped:
                        prev_lower = deduped[-1].lower().replace(" ", "_")
                        p_lower = p.lower().replace(" ", "_")
                        if p_lower.startswith(prev_lower + "_"):
                            p = p[len(deduped[-1]) + 1:]
                    if p and (not deduped or deduped[-1] != p):
                        deduped.append(p)
                composite = "_".join(deduped)
                # Use _normalize_key but strip the col_ prefix for temporal columns
                # since digit-leading names like "2015_q1_jan" are expected
                key = _normalize_key(composite)
                if key.startswith("col_"):
                    key = key[4:]
                headers.append(key)
            else:
                headers.append(f"col_{col_idx}")

    # Build data records, filtering out rows where all data columns are null.
    # Forward-fill label columns so grouped rows inherit their parent label.
    last_labels: list[Any] = [None] * label_cols
    result: list[dict[str, Any]] = []
    for row in data_rows:
        # Skip rows where all values are None
        if all(v is None for v in row):
            continue
        # Skip rows where all DATA columns (after label cols) are None
        data_values = list(row[label_cols:])
        if all(v is None for v in data_values):
            continue
        record: dict[str, Any] = {}
        for i, v in enumerate(row):
            if i >= len(headers):
                break
            if i < label_cols:
                # Forward-fill: use current value if present, else last seen
                if v is not None:
                    last_labels[i] = v
                record[headers[i]] = _mask_privacy(last_labels[i])
            else:
                record[headers[i]] = _mask_privacy(v)
        result.append(record)

    return result


def _parse_csv(content: bytes, skip_rows: int = 0) -> list[dict[str, Any]]:
    """Parse CSV bytes (including BOM-prefixed) into a list of dicts.

    Args:
        content: Raw CSV file bytes (UTF-8 or UTF-8-BOM).
        skip_rows: Number of lines to skip before the header row.

    Returns:
        list of dicts with normalized snake_case keys and privacy-masked values.
    """
    text = content.decode("utf-8-sig")
    lines = text.splitlines(keepends=True)
    lines = lines[skip_rows:]
    reader = csv.DictReader(StringIO("".join(lines)))
    result: list[dict[str, Any]] = []
    for row in reader:
        normalized = {_normalize_key(k): _mask_privacy(v) for k, v in row.items()}
        result.append(normalized)
    return result


async def fetch_and_parse(
    url: str,
    sheet: str | int = 0,
    skip_rows: int = 0,
    ttl: int = 86400,
    ircc_parse_config: dict | None = None,
) -> tuple[list[dict[str, Any]], bool]:
    """Fetch a remote file and parse it into a list of dicts.

    Supports XLSX, XLS, and CSV files. Results are cached via cached_fetch()
    with the given TTL. Exceptions from the HTTP fetch or parsing are NOT
    cached — they propagate to the caller.

    Args:
        url: Remote file URL. Routing by suffix: .csv -> CSV, .xls -> XLS,
             anything else (including .xlsx) -> XLSX.
        sheet: Sheet name or 0-based index (XLSX/XLS only).
        skip_rows: Rows to skip before the header row.
        ttl: Cache TTL in seconds (default: 86400 = 24 hours).
        ircc_parse_config: When provided (dict with skip_rows, header_rows, label_cols),
            routes XLSX files through _parse_ircc_xlsx for multi-row merged header support.
            Non-IRCC callers that omit this parameter get identical existing behavior.

    Returns:
        (list[dict], was_cached) where was_cached=True if served from cache.
    """
    config_hash = str(sorted(ircc_parse_config.items())) if ircc_parse_config else ""
    cache_key = f"parsers:{url}:{sheet}:{skip_rows}:{config_hash}"

    async def _fetch() -> list[dict[str, Any]]:
        async with httpx.AsyncClient(timeout=60.0) as client:
            response = await client.get(url)
            response.raise_for_status()
            raw = response.content

        lower_url = url.lower().split("?")[0]
        if lower_url.endswith(".csv"):
            return _parse_csv(raw, skip_rows)
        elif lower_url.endswith(".xls"):
            return _parse_xls(raw, sheet, skip_rows)
        elif ircc_parse_config is not None:
            return _parse_ircc_xlsx(raw, sheet=sheet, **ircc_parse_config)
        else:
            return _parse_xlsx(raw, sheet, skip_rows)

    return await cached_fetch(cache_key, ttl, _fetch)
