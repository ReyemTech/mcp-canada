"""Unit tests for shared/parsers.py."""

from __future__ import annotations

import csv
import io
from io import BytesIO
from typing import Any
from unittest.mock import AsyncMock, MagicMock, patch

import openpyxl
import pytest


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_xlsx_bytes(
    headers: list[str],
    rows: list[list[Any]],
    sheet_name: str = "Sheet1",
) -> bytes:
    """Create minimal XLSX bytes using openpyxl for test fixtures."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_csv_bytes(headers: list[str], rows: list[list[Any]], bom: bool = False) -> bytes:
    """Create CSV bytes for test fixtures."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    text = buf.getvalue()
    if bom:
        text = "\ufeff" + text
    return text.encode("utf-8")


# ---------------------------------------------------------------------------
# _normalize_key
# ---------------------------------------------------------------------------


class TestNormalizeKey:
    def test_simple_header(self) -> None:
        from mcp_canada.shared.parsers import _normalize_key

        assert _normalize_key("Country of Citizenship") == "country_of_citizenship"

    def test_strips_whitespace(self) -> None:
        from mcp_canada.shared.parsers import _normalize_key

        assert _normalize_key("  Year ") == "year"

    def test_leading_digit_prefixed(self) -> None:
        from mcp_canada.shared.parsers import _normalize_key

        assert _normalize_key("123col") == "col_123col"

    def test_empty_returns_col(self) -> None:
        from mcp_canada.shared.parsers import _normalize_key

        assert _normalize_key("") == "col"

    def test_special_chars_to_underscore(self) -> None:
        from mcp_canada.shared.parsers import _normalize_key

        assert _normalize_key("a/b-c d") == "a_b_c_d"

    def test_multiple_underscores_collapsed(self) -> None:
        from mcp_canada.shared.parsers import _normalize_key

        assert _normalize_key("a  b") == "a_b"


# ---------------------------------------------------------------------------
# _mask_privacy
# ---------------------------------------------------------------------------


class TestMaskPrivacy:
    def test_double_dash_returns_none(self) -> None:
        from mcp_canada.shared.parsers import _mask_privacy

        assert _mask_privacy("--") is None

    def test_double_dash_with_whitespace_returns_none(self) -> None:
        from mcp_canada.shared.parsers import _mask_privacy

        assert _mask_privacy(" -- ") is None

    def test_numeric_string_converted_to_int(self) -> None:
        from mcp_canada.shared.parsers import _mask_privacy

        assert _mask_privacy("123") == 123
        assert _mask_privacy("2,630") == 2630
        assert _mask_privacy("  13 365") == 13365
        assert _mask_privacy("1.5") == 1.5

    def test_int_unchanged(self) -> None:
        from mcp_canada.shared.parsers import _mask_privacy

        assert _mask_privacy(42) == 42

    def test_none_returns_none(self) -> None:
        from mcp_canada.shared.parsers import _mask_privacy

        assert _mask_privacy(None) is None


# ---------------------------------------------------------------------------
# _parse_xlsx via pandas path
# ---------------------------------------------------------------------------


_has_pandas = True
try:
    import pandas  # noqa: F401
except ImportError:
    _has_pandas = False


@pytest.mark.skipif(not _has_pandas, reason="pandas not installed")
class TestParseXlsxPandasPath:
    def test_uses_pandas_when_importable(self) -> None:
        """_parse_xlsx() calls pandas.read_excel when pandas is available."""
        import pandas as pd

        from mcp_canada.shared.parsers import _parse_xlsx

        content = _make_xlsx_bytes(["Year", "Count"], [[2020, 100]])
        df = pd.DataFrame([{"Year": 2020, "Count": 100}])

        with patch("pandas.read_excel", return_value=df) as mock_read:
            result = _parse_xlsx(content, sheet=0, skip_rows=0)

        mock_read.assert_called_once()
        call_args = mock_read.call_args
        # First positional arg should be a BytesIO
        assert isinstance(call_args[0][0], BytesIO)
        assert call_args[1]["sheet_name"] == 0
        assert call_args[1]["skiprows"] == 0
        assert isinstance(result, list)

    def test_parse_xlsx_pandas_normalized_keys(self) -> None:
        """_parse_xlsx_pandas returns list[dict] with normalized keys."""
        import pandas as pd

        from mcp_canada.shared.parsers import _parse_xlsx_pandas

        content = _make_xlsx_bytes(
            ["Country of Citizenship", "Total"],
            [["Canada", 1000], ["USA", "--"]],
        )
        df = pd.DataFrame(
            [
                {"Country of Citizenship": "Canada", "Total": 1000},
                {"Country of Citizenship": "USA", "Total": "--"},
            ]
        )
        with patch("pandas.read_excel", return_value=df):
            result = _parse_xlsx_pandas(content)

        assert len(result) == 2
        assert "country_of_citizenship" in result[0]
        assert "total" in result[0]
        assert result[1]["total"] is None  # '--' masked


# ---------------------------------------------------------------------------
# _parse_xlsx via openpyxl fallback
# ---------------------------------------------------------------------------


class TestParseXlsxOpenpyxlPath:
    def test_falls_back_to_openpyxl_on_pandas_import_error(self) -> None:
        """_parse_xlsx() uses openpyxl when pandas import fails."""
        content = _make_xlsx_bytes(["Year", "Count"], [[2020, 100]])

        import builtins

        real_import = builtins.__import__

        def fake_import(name: str, *args: Any, **kwargs: Any) -> Any:
            if name == "pandas":
                raise ImportError("no pandas")
            return real_import(name, *args, **kwargs)

        with patch("builtins.__import__", side_effect=fake_import):
            from mcp_canada.shared import parsers as parsers_mod

            # Patch _parse_xlsx_openpyxl on the module to verify it's called
            with patch.object(parsers_mod, "_parse_xlsx_openpyxl", return_value=[]) as mock_oxl:
                parsers_mod._parse_xlsx(content, sheet=0, skip_rows=0)
                mock_oxl.assert_called_once_with(content, 0, 0)

    def test_parse_xlsx_openpyxl_normalized_keys(self) -> None:
        """_parse_xlsx_openpyxl returns list[dict] with normalized keys and masked values."""
        from mcp_canada.shared.parsers import _parse_xlsx_openpyxl

        content = _make_xlsx_bytes(
            ["Country of Citizenship", "Total"],
            [["Canada", 1000], ["USA", "--"]],
        )
        result = _parse_xlsx_openpyxl(content)

        assert len(result) == 2
        assert result[0]["country_of_citizenship"] == "Canada"
        assert result[0]["total"] == 1000
        assert result[1]["total"] is None  # '--' masked

    def test_parse_xlsx_openpyxl_skip_rows(self) -> None:
        """_parse_xlsx_openpyxl with skip_rows=1 skips first row before header."""
        from mcp_canada.shared.parsers import _parse_xlsx_openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["metadata row"])
        ws.append(["Year", "Count"])
        ws.append([2020, 100])
        buf = BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        result = _parse_xlsx_openpyxl(content, skip_rows=1)

        assert len(result) == 1
        assert "year" in result[0]
        assert result[0]["year"] == 2020

    def test_parse_xlsx_openpyxl_sheet_by_index(self) -> None:
        """_parse_xlsx_openpyxl selects the correct sheet by index."""
        from mcp_canada.shared.parsers import _parse_xlsx_openpyxl

        wb = openpyxl.Workbook()
        ws0 = wb.active
        assert ws0 is not None
        ws0.title = "First"
        ws0.append(["A"])
        ws0.append([1])
        ws1 = wb.create_sheet("Second")
        ws1.append(["B"])
        ws1.append([2])
        buf = BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        result = _parse_xlsx_openpyxl(content, sheet=1)
        assert result[0]["b"] == 2

    def test_parse_xlsx_openpyxl_empty_workbook(self) -> None:
        """_parse_xlsx_openpyxl with empty workbook returns []."""
        from mcp_canada.shared.parsers import _parse_xlsx_openpyxl

        wb = openpyxl.Workbook()
        buf = BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        result = _parse_xlsx_openpyxl(content)
        assert result == []


# ---------------------------------------------------------------------------
# _parse_csv
# ---------------------------------------------------------------------------


class TestParseCsv:
    def test_bom_prefixed_csv_normalized_keys(self) -> None:
        """_parse_csv with BOM-prefixed CSV returns list[dict] with normalized keys."""
        from mcp_canada.shared.parsers import _parse_csv

        content = _make_csv_bytes(
            ["Country of Citizenship", "Year"],
            [["Canada", 2020]],
            bom=True,
        )
        result = _parse_csv(content)

        assert len(result) == 1
        assert "country_of_citizenship" in result[0]
        assert result[0]["year"] == 2020

    def test_privacy_masking_in_csv(self) -> None:
        """_parse_csv with '--' values returns None in output."""
        from mcp_canada.shared.parsers import _parse_csv

        content = _make_csv_bytes(
            ["Name", "Count"],
            [["Alice", "--"], ["Bob", "5"]],
        )
        result = _parse_csv(content)

        assert result[0]["count"] is None
        assert result[1]["count"] == 5


# ---------------------------------------------------------------------------
# fetch_and_parse
# ---------------------------------------------------------------------------


class TestFetchAndParse:
    @pytest.mark.asyncio
    async def test_calls_cached_fetch_and_returns_tuple(self) -> None:
        """fetch_and_parse calls cached_fetch with correct key and returns (list[dict], bool)."""
        fake_data = [{"year": 2020}]

        with patch(
            "mcp_canada.shared.parsers.cached_fetch",
            new_callable=AsyncMock,
            return_value=(fake_data, True),
        ) as mock_cf:
            from mcp_canada.shared.parsers import fetch_and_parse

            result, was_cached = await fetch_and_parse(
                "https://example.com/data.xlsx", sheet=0, skip_rows=0, ttl=86400
            )

        mock_cf.assert_called_once()
        call_key = mock_cf.call_args[0][0]
        assert "data.xlsx" in call_key
        assert result == fake_data
        assert was_cached is True

    @pytest.mark.asyncio
    async def test_routes_csv_to_parse_csv(self) -> None:
        """fetch_and_parse detects .csv suffix and routes to _parse_csv."""
        csv_bytes = _make_csv_bytes(["Year"], [[2020]])

        async def fake_cached_fetch(key: str, ttl: int, fetcher: Any) -> tuple[Any, bool]:
            data = await fetcher()
            return data, False

        mock_response = MagicMock()
        mock_response.content = csv_bytes
        mock_response.raise_for_status = MagicMock()

        with (
            patch("mcp_canada.shared.parsers.cached_fetch", side_effect=fake_cached_fetch),
            patch("httpx.AsyncClient") as mock_client_cls,
            patch("mcp_canada.shared.parsers._parse_csv", return_value=[{"year": "2020"}]) as mock_csv,
        ):
            mock_client = AsyncMock()
            mock_client.__aenter__ = AsyncMock(return_value=mock_client)
            mock_client.__aexit__ = AsyncMock(return_value=None)
            mock_client.get = AsyncMock(return_value=mock_response)
            mock_client_cls.return_value = mock_client

            from mcp_canada.shared.parsers import fetch_and_parse

            result, _ = await fetch_and_parse("https://example.com/data.csv")

        mock_csv.assert_called_once_with(csv_bytes, 0)

    @pytest.mark.asyncio
    async def test_routes_xlsx_to_parse_xlsx(self) -> None:
        """fetch_and_parse detects .xlsx suffix and routes to _parse_xlsx."""
        xlsx_bytes = _make_xlsx_bytes(["Year"], [[2020]])

        async def fake_cached_fetch(key: str, ttl: int, fetcher: Any) -> tuple[Any, bool]:
            data = await fetcher()
            return data, False

        mock_response = MagicMock()
        mock_response.content = xlsx_bytes
        mock_response.raise_for_status = MagicMock()

        with (
            patch("mcp_canada.shared.parsers.cached_fetch", side_effect=fake_cached_fetch),
            patch("httpx.AsyncClient") as mock_client_cls,
            patch("mcp_canada.shared.parsers._parse_xlsx", return_value=[{"year": 2020}]) as mock_xlsx,
        ):
            mock_client = AsyncMock()
            mock_client.__aenter__ = AsyncMock(return_value=mock_client)
            mock_client.__aexit__ = AsyncMock(return_value=None)
            mock_client.get = AsyncMock(return_value=mock_response)
            mock_client_cls.return_value = mock_client

            from mcp_canada.shared.parsers import fetch_and_parse

            result, _ = await fetch_and_parse("https://example.com/data.xlsx")

        mock_xlsx.assert_called_once_with(xlsx_bytes, 0, 0)

    @pytest.mark.asyncio
    async def test_routes_wfs_csv_via_query_param(self) -> None:
        """MTQ WFS URL with ?outputformat=csv routes to _parse_csv even though path has no .csv suffix."""
        csv_bytes = _make_csv_bytes(["Col"], [["val"]])

        async def fake_cached_fetch(key: str, ttl: int, fetcher: Any) -> tuple[Any, bool]:
            data = await fetcher()
            return data, False

        mock_response = MagicMock()
        mock_response.content = csv_bytes
        mock_response.raise_for_status = MagicMock()

        with (
            patch("mcp_canada.shared.parsers.cached_fetch", side_effect=fake_cached_fetch),
            patch("httpx.AsyncClient") as mock_client_cls,
            patch("mcp_canada.shared.parsers._parse_csv", return_value=[{"col": "val"}]) as mock_csv,
            patch("mcp_canada.shared.parsers._parse_xlsx") as mock_xlsx,
        ):
            mock_client = AsyncMock()
            mock_client.__aenter__ = AsyncMock(return_value=mock_client)
            mock_client.__aexit__ = AsyncMock(return_value=None)
            mock_client.get = AsyncMock(return_value=mock_response)
            mock_client_cls.return_value = mock_client

            from mcp_canada.shared.parsers import fetch_and_parse

            url = (
                "https://ws.mapserver.transports.gouv.qc.ca/swtq"
                "?service=wfs&version=2.0.0&request=getfeature"
                "&typename=ms:gsq_v_desc_strct_tri&outputformat=csv"
            )
            result, _ = await fetch_and_parse(url)

        mock_csv.assert_called_once_with(csv_bytes, 0)
        mock_xlsx.assert_not_called()
        assert result == [{"col": "val"}]

    @pytest.mark.asyncio
    async def test_routes_wfs_csv_via_mixed_case_outputFormat(self) -> None:
        """?outputFormat=csv (mixed case key) routes to _parse_csv — key comparison is case-insensitive."""
        csv_bytes = _make_csv_bytes(["A"], [["b"]])

        async def fake_cached_fetch(key: str, ttl: int, fetcher: Any) -> tuple[Any, bool]:
            return await fetcher(), False

        mock_response = MagicMock()
        mock_response.content = csv_bytes
        mock_response.raise_for_status = MagicMock()

        with (
            patch("mcp_canada.shared.parsers.cached_fetch", side_effect=fake_cached_fetch),
            patch("httpx.AsyncClient") as mock_client_cls,
            patch("mcp_canada.shared.parsers._parse_csv", return_value=[{"a": "b"}]) as mock_csv,
        ):
            mock_client = AsyncMock()
            mock_client.__aenter__ = AsyncMock(return_value=mock_client)
            mock_client.__aexit__ = AsyncMock(return_value=None)
            mock_client.get = AsyncMock(return_value=mock_response)
            mock_client_cls.return_value = mock_client

            from mcp_canada.shared.parsers import fetch_and_parse

            result, _ = await fetch_and_parse(
                "https://ws.example.com/swtq?typeName=ms:roads&outputFormat=csv"
            )

        mock_csv.assert_called_once()
        assert result == [{"a": "b"}]

    @pytest.mark.asyncio
    async def test_routes_wfs_csv_via_format_key(self) -> None:
        """?format=csv short key also routes to _parse_csv."""
        csv_bytes = _make_csv_bytes(["X"], [["y"]])

        async def fake_cached_fetch(key: str, ttl: int, fetcher: Any) -> tuple[Any, bool]:
            return await fetcher(), False

        mock_response = MagicMock()
        mock_response.content = csv_bytes
        mock_response.raise_for_status = MagicMock()

        with (
            patch("mcp_canada.shared.parsers.cached_fetch", side_effect=fake_cached_fetch),
            patch("httpx.AsyncClient") as mock_client_cls,
            patch("mcp_canada.shared.parsers._parse_csv", return_value=[{"x": "y"}]) as mock_csv,
        ):
            mock_client = AsyncMock()
            mock_client.__aenter__ = AsyncMock(return_value=mock_client)
            mock_client.__aexit__ = AsyncMock(return_value=None)
            mock_client.get = AsyncMock(return_value=mock_response)
            mock_client_cls.return_value = mock_client

            from mcp_canada.shared.parsers import fetch_and_parse

            await fetch_and_parse("https://api.example.com/wfs?format=csv")

        mock_csv.assert_called_once()

    @pytest.mark.asyncio
    async def test_path_suffix_still_wins_for_xlsx_even_with_irrelevant_query(self) -> None:
        """Backward compat: .xlsx path suffix routes to _parse_xlsx regardless of unrelated query params."""
        xlsx_bytes = _make_xlsx_bytes(["Year"], [[2020]])

        async def fake_cached_fetch(key: str, ttl: int, fetcher: Any) -> tuple[Any, bool]:
            return await fetcher(), False

        mock_response = MagicMock()
        mock_response.content = xlsx_bytes
        mock_response.raise_for_status = MagicMock()

        with (
            patch("mcp_canada.shared.parsers.cached_fetch", side_effect=fake_cached_fetch),
            patch("httpx.AsyncClient") as mock_client_cls,
            patch("mcp_canada.shared.parsers._parse_xlsx", return_value=[{"year": 2020}]) as mock_xlsx,
            patch("mcp_canada.shared.parsers._parse_csv") as mock_csv,
        ):
            mock_client = AsyncMock()
            mock_client.__aenter__ = AsyncMock(return_value=mock_client)
            mock_client.__aexit__ = AsyncMock(return_value=None)
            mock_client.get = AsyncMock(return_value=mock_response)
            mock_client_cls.return_value = mock_client

            from mcp_canada.shared.parsers import fetch_and_parse

            await fetch_and_parse("https://example.com/data.xlsx?cacheBust=123")

        mock_xlsx.assert_called_once()
        mock_csv.assert_not_called()

    @pytest.mark.asyncio
    async def test_no_format_hint_falls_through_to_xlsx(self) -> None:
        """URL with neither suffix nor format query key still falls through to _parse_xlsx (existing behaviour)."""
        xlsx_bytes = _make_xlsx_bytes(["A"], [["b"]])

        async def fake_cached_fetch(key: str, ttl: int, fetcher: Any) -> tuple[Any, bool]:
            return await fetcher(), False

        mock_response = MagicMock()
        mock_response.content = xlsx_bytes
        mock_response.raise_for_status = MagicMock()

        with (
            patch("mcp_canada.shared.parsers.cached_fetch", side_effect=fake_cached_fetch),
            patch("httpx.AsyncClient") as mock_client_cls,
            patch("mcp_canada.shared.parsers._parse_xlsx", return_value=[{"a": "b"}]) as mock_xlsx,
        ):
            mock_client = AsyncMock()
            mock_client.__aenter__ = AsyncMock(return_value=mock_client)
            mock_client.__aexit__ = AsyncMock(return_value=None)
            mock_client.get = AsyncMock(return_value=mock_response)
            mock_client_cls.return_value = mock_client

            from mcp_canada.shared.parsers import fetch_and_parse

            await fetch_and_parse("https://api.example.com/resource?id=123&token=abc")

        mock_xlsx.assert_called_once()


# ---------------------------------------------------------------------------
# _parse_ircc_xlsx
# ---------------------------------------------------------------------------


def _make_ircc_xlsx_layout_a(
    years: list[int],
    quarters: list[str],
    months: list[str],
    data_rows: list[list],
) -> bytes:
    """Create a Layout A IRCC XLSX fixture.

    Structure:
      Row 1: Title (merged across all columns)
      Row 2: Blank (merged)
      Row 3: "Country of Citizenship" | year1 | year1 | ... | yearN (merged per year)
      Row 4: empty | Q1 | Q1 | Q1 | Q1 | Year Total | Q2 | ... (merged per quarter)
      Row 5: empty | Jan | Feb | Mar | Q1 Total | Apr | ... | QN Total | Year Total (no merges)
      Row 6+: Data rows — label (col A) + values
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None

    # Calculate total columns: 1 label + len(years) * (len(months)/year + 1 q_total columns + 1 year_total)
    # For simplicity, use small fixture: 1 year, 2 quarters, 2 months each + totals
    # Build headers based on provided args
    # For tests we use: 1 label col + columns per combination
    # Build a simple 1 year, 1 quarter, 2 months structure + totals
    # Actual structure: label | 2015 (Q1 | Jan | Feb | Q1_Total | Year_Total)
    # That is: 1 label + (len(months) + 1 q_total + 1 year_total) per quarter per year

    # Simple fixture: 1 year (2015), 1 quarter (Q1), 2 months (Jan, Feb) + Q1 Total + Year Total
    # Total cols = 1 + 2 + 1 + 1 = 5
    # Row 1: Title merged A1:E1
    # Row 2: Blank
    # Row 3: "Country of Citizenship" in A | "2015" merged B3:E3
    # Row 4: blank in A | "Q1" merged B4:D4 | "Year Total" in E4
    # Row 5: blank in A | "Jan" in B | "Feb" in C | "Q1 Total" in D | blank in E
    # Row 6+: data

    n_months_per_q = 2
    total_data_cols = n_months_per_q + 1  # months + Q total
    total_cols = 1 + total_data_cols + 1  # label + (months + q_total) + year_total

    # Row 1: Title
    ws.append(["IRCC Data Title"] + [None] * (total_cols - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # Row 2: Blank
    ws.append([None] * total_cols)

    # Row 3: Label header + Year merges
    row3 = ["Country of Citizenship", "2015", None, None, None]
    ws.append(row3)
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=total_cols)

    # Row 4: blank + Quarter merges + Year Total
    row4 = [None, "Q1", None, None, "Year Total"]
    ws.append(row4)
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=4)

    # Row 5: blank + Month labels + Q Total
    row5 = [None, "Jan", "Feb", "Q1 Total", None]
    ws.append(row5)

    # Data rows
    for data_row in data_rows:
        ws.append(data_row)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_ircc_xlsx_layout_b(data_rows: list[list]) -> bytes:
    """Create a Layout B IRCC XLSX fixture with 2 label columns, 3 header rows."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None

    # Total cols: 2 labels + Jan + Feb + Q1 Total + Year Total = 6
    total_cols = 6

    # Row 1: Title merged
    ws.append(["IRCC EE Data"] + [None] * (total_cols - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # Row 2: Blank
    ws.append([None] * total_cols)

    # Row 3: label header merged A3:B3 + Year 2015 merged C3:F3
    row3 = ["Gender and Province", None, "2015", None, None, None]
    ws.append(row3)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=6)

    # Row 4: blank A4:B4 + Q1 merged C4:E4 + Year Total F4
    row4 = [None, None, "Q1", None, None, "Year Total"]
    ws.append(row4)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)
    ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=5)

    # Row 5: blank + blank + Jan + Feb + Q1 Total + blank
    row5 = [None, None, "Jan", "Feb", "Q1 Total", None]
    ws.append(row5)

    # Data rows
    for data_row in data_rows:
        ws.append(data_row)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_ircc_xlsx_layout_b_monthly(data_rows: list[list]) -> bytes:
    """Create a Layout B-monthly IRCC XLSX fixture: 2 labels, 2 header rows (no quarter row)."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None

    # Total cols: 2 labels + Jan + Feb + Year Total = 5
    total_cols = 5

    # Row 1: Title merged
    ws.append(["IRCC Asylum Data"] + [None] * (total_cols - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # Row 2: Blank
    ws.append([None] * total_cols)

    # Row 3: label merged A3:B3 + Year 2015 merged C3:E3
    row3 = ["Office Type and Province", None, "2015", None, None]
    ws.append(row3)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=5)

    # Row 4: blank A4:B4 + Jan + Feb + Year Total
    row4 = [None, None, "Jan", "Feb", "Year Total"]
    ws.append(row4)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)

    # Data rows
    for data_row in data_rows:
        ws.append(data_row)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_ircc_xlsx_layout_c(data_rows: list[list]) -> bytes:
    """Create a Layout C IRCC XLSX fixture: ops layout (6 skip rows, 2 header rows)."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None

    # Total cols: 1 label + January + February + Year Total = 4
    total_cols = 4

    # Rows 1-3: Empty
    for _ in range(3):
        ws.append([None] * total_cols)

    # Row 4: Title merged
    ws.append(["PR Intake Operations Data"] + [None] * (total_cols - 1))
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=total_cols)

    # Rows 5-6: Empty
    for _ in range(2):
        ws.append([None] * total_cols)

    # Row 7 (skip_rows=6): Label col header merged A7:A8 + Year 2023 merged B7:D7
    row7 = ["Application Type", "2023", None, None]
    ws.append(row7)
    ws.merge_cells(start_row=7, start_column=1, end_row=8, end_column=1)
    ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=4)

    # Row 8: blank (label merged from A7) + Month labels with trailing spaces
    row8 = [None, "January  ", "February ", "Year Total"]
    ws.append(row8)

    # Data rows
    for data_row in data_rows:
        ws.append(data_row)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestParseIrccXlsx:
    """Tests for _parse_ircc_xlsx: IRCC multi-row merged header parsing."""

    def test_layout_a_composite_column_names(self) -> None:
        """Layout A (1 label col, 3 header rows): produces composite col names."""
        from mcp_canada.shared.parsers import _parse_ircc_xlsx

        content = _make_ircc_xlsx_layout_a(
            years=[2015], quarters=["Q1"], months=["Jan", "Feb"],
            data_rows=[["Canada", "90", "2,630", "2,720", "2,720"]],
        )
        result = _parse_ircc_xlsx(content, skip_rows=2, header_rows=3, label_cols=1)

        assert len(result) == 1
        row = result[0]
        # Label col: first header row value
        assert "country_of_citizenship" in row
        # Composite temporal cols: year_quarter_month
        assert "2015_q1_jan" in row
        assert "2015_q1_feb" in row
        assert "2015_q1_total" in row
        assert "2015_year_total" in row

    def test_layout_a_data_values_returned_as_is(self) -> None:
        """Layout A: numeric string values like '2,630' and '90' returned unchanged."""
        from mcp_canada.shared.parsers import _parse_ircc_xlsx

        content = _make_ircc_xlsx_layout_a(
            years=[2015], quarters=["Q1"], months=["Jan", "Feb"],
            data_rows=[["Canada", "90", "2,630", "2,720", "2,720"]],
        )
        result = _parse_ircc_xlsx(content, skip_rows=2, header_rows=3, label_cols=1)

        row = result[0]
        assert row["country_of_citizenship"] == "Canada"
        assert row["2015_q1_jan"] == 90
        assert row["2015_q1_feb"] == 2630

    def test_layout_a_privacy_masking(self) -> None:
        """Layout A: '--' values are masked to None."""
        from mcp_canada.shared.parsers import _parse_ircc_xlsx

        content = _make_ircc_xlsx_layout_a(
            years=[2015], quarters=["Q1"], months=["Jan", "Feb"],
            data_rows=[["Country X", "--", "100", "100", "100"]],
        )
        result = _parse_ircc_xlsx(content, skip_rows=2, header_rows=3, label_cols=1)

        assert result[0]["2015_q1_jan"] is None
        assert result[0]["2015_q1_feb"] == 100

    def test_layout_a_strips_trailing_empty_rows(self) -> None:
        """Layout A: all-None data rows at end are stripped."""
        from mcp_canada.shared.parsers import _parse_ircc_xlsx

        content = _make_ircc_xlsx_layout_a(
            years=[2015], quarters=["Q1"], months=["Jan", "Feb"],
            data_rows=[
                ["Canada", "100", "200", "300", "300"],
                [None, None, None, None, None],  # trailing empty
            ],
        )
        result = _parse_ircc_xlsx(content, skip_rows=2, header_rows=3, label_cols=1)

        assert len(result) == 1

    def test_layout_b_two_label_cols(self) -> None:
        """Layout B (2 label cols, 3 header rows): produces 2 label columns."""
        from mcp_canada.shared.parsers import _parse_ircc_xlsx

        content = _make_ircc_xlsx_layout_b(
            data_rows=[["Male", "Ontario", "50", "60", "110", "110"]],
        )
        result = _parse_ircc_xlsx(content, skip_rows=2, header_rows=3, label_cols=2)

        assert len(result) == 1
        row = result[0]
        # Should have 2 label columns derived from the merged header
        keys = list(row.keys())
        # First 2 keys are label cols
        assert row[keys[0]] == "Male"
        assert row[keys[1]] == "Ontario"
        # Temporal col names exist
        assert any("2015" in k for k in row.keys())

    def test_layout_b_monthly_no_quarter_row(self) -> None:
        """Layout B-monthly (2 header rows, no quarter): produces year_month columns."""
        from mcp_canada.shared.parsers import _parse_ircc_xlsx

        content = _make_ircc_xlsx_layout_b_monthly(
            data_rows=[["Airport", "Toronto", "10", "20", "30"]],
        )
        result = _parse_ircc_xlsx(content, skip_rows=2, header_rows=2, label_cols=2)

        assert len(result) == 1
        row = result[0]
        # Should have col_2015_jan (no quarter segment)
        assert "2015_jan" in row
        assert "2015_feb" in row

    def test_layout_c_ops_year_month_columns(self) -> None:
        """Layout C (ops, 6 skip rows, 2 header rows): produces year_month columns."""
        from mcp_canada.shared.parsers import _parse_ircc_xlsx

        content = _make_ircc_xlsx_layout_c(
            data_rows=[["PR Intake", 100, 200, 300]],
        )
        result = _parse_ircc_xlsx(content, skip_rows=6, header_rows=2, label_cols=1)

        assert len(result) == 1
        row = result[0]
        assert "2023_january" in row
        assert "2023_february" in row

    def test_merged_cells_forward_filled(self) -> None:
        """Merged header cells produce None in non-anchor positions; must be forward-filled."""
        from mcp_canada.shared.parsers import _parse_ircc_xlsx

        # Layout A: year "2015" is only in col B3; C3, D3, E3 are None due to merge
        # After forward-fill, all 4 positions should have "2015"
        content = _make_ircc_xlsx_layout_a(
            years=[2015], quarters=["Q1"], months=["Jan", "Feb"],
            data_rows=[["Canada", "1", "2", "3", "4"]],
        )
        result = _parse_ircc_xlsx(content, skip_rows=2, header_rows=3, label_cols=1)

        row = result[0]
        # All temporal columns should have "2015" as the year prefix
        temporal_keys = [k for k in row.keys() if k.startswith("2015")]
        assert len(temporal_keys) == 4  # Jan, Feb, Q1 Total, Year Total

    def test_existing_parse_xlsx_unchanged(self) -> None:
        """Existing _parse_xlsx behavior is unchanged (backward compat check)."""
        from mcp_canada.shared.parsers import _parse_xlsx

        content = _make_xlsx_bytes(
            ["Country", "Total"],
            [["Canada", 1000], ["USA", 500]],
        )
        result = _parse_xlsx(content)

        assert len(result) == 2
        assert result[0]["country"] == "Canada"
        assert result[0]["total"] == 1000

    def test_fetch_and_parse_ircc_config_kwarg(self) -> None:
        """fetch_and_parse accepts ircc_parse_config kwarg without error."""
        # This verifies the signature extension doesn't break existing callers
        import inspect

        from mcp_canada.shared.parsers import fetch_and_parse

        sig = inspect.signature(fetch_and_parse)
        assert "ircc_parse_config" in sig.parameters


# ---------------------------------------------------------------------------
# _parse_geojson
# ---------------------------------------------------------------------------


import json  # noqa: E402


class TestGeoJSON:
    def _make_feature_collection(
        self, features: list[dict[str, Any]]
    ) -> bytes:
        """Build GeoJSON FeatureCollection bytes."""
        fc = {"type": "FeatureCollection", "features": features}
        return json.dumps(fc).encode("utf-8")

    def _make_feature(
        self,
        properties: dict[str, Any] | None,
        geometry: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        if geometry is None:
            geometry = {"type": "Point", "coordinates": [-79.38, 43.65]}
        return {
            "type": "Feature",
            "geometry": geometry,
            "properties": properties,
        }

    def test_three_features_returns_three_dicts(self) -> None:
        """FeatureCollection with 3 features returns 3 property dicts."""
        from mcp_canada.shared.parsers import _parse_geojson

        features = [
            self._make_feature({"name": "A", "value": 1}),
            self._make_feature({"name": "B", "value": 2}),
            self._make_feature({"name": "C", "value": 3}),
        ]
        content = self._make_feature_collection(features)
        result = _parse_geojson(content)

        assert len(result) == 3
        assert result[0] == {"name": "A", "value": 1}
        assert result[1] == {"name": "B", "value": 2}
        assert result[2] == {"name": "C", "value": 3}

    def test_no_geometry_key_by_default(self) -> None:
        """By default, geometry key is NOT included in output dicts."""
        from mcp_canada.shared.parsers import _parse_geojson

        features = [self._make_feature({"name": "A"})]
        content = self._make_feature_collection(features)
        result = _parse_geojson(content)

        assert "geometry" not in result[0]

    def test_include_geometry_true_adds_geometry_key(self) -> None:
        """With include_geometry=True, each dict includes the 'geometry' key."""
        from mcp_canada.shared.parsers import _parse_geojson

        geom = {"type": "Point", "coordinates": [-79.38, 43.65]}
        features = [self._make_feature({"name": "A"}, geometry=geom)]
        content = self._make_feature_collection(features)
        result = _parse_geojson(content, include_geometry=True)

        assert "geometry" in result[0]
        assert result[0]["geometry"] == geom

    def test_empty_feature_collection_returns_empty_list(self) -> None:
        """Empty FeatureCollection returns []."""
        from mcp_canada.shared.parsers import _parse_geojson

        content = self._make_feature_collection([])
        result = _parse_geojson(content)

        assert result == []

    def test_feature_without_properties_returns_empty_dict(self) -> None:
        """Feature with null properties returns an empty dict {}."""
        from mcp_canada.shared.parsers import _parse_geojson

        feature = {"type": "Feature", "geometry": None, "properties": None}
        content = self._make_feature_collection([feature])
        result = _parse_geojson(content)

        assert result == [{}]


# ---------------------------------------------------------------------------
# _parse_json
# ---------------------------------------------------------------------------


class TestParseJSON:
    def test_json_array_returns_list_of_dicts(self) -> None:
        """JSON array of dicts returns list[dict] directly."""
        from mcp_canada.shared.parsers import _parse_json

        data = [{"id": 1, "name": "Alice"}, {"id": 2, "name": "Bob"}]
        content = json.dumps(data).encode("utf-8")
        result = _parse_json(content)

        assert result == data

    def test_json_object_with_features_delegates_to_geojson(self) -> None:
        """JSON object with 'features' key delegates to _parse_geojson."""
        from mcp_canada.shared.parsers import _parse_json

        fc = {
            "type": "FeatureCollection",
            "features": [
                {"type": "Feature", "geometry": None, "properties": {"x": 1}},
            ],
        }
        content = json.dumps(fc).encode("utf-8")
        result = _parse_json(content)

        assert result == [{"x": 1}]

    def test_plain_json_object_wrapped_in_list(self) -> None:
        """Plain JSON object (no 'features' key) is wrapped in a list."""
        from mcp_canada.shared.parsers import _parse_json

        obj = {"key": "value", "count": 42}
        content = json.dumps(obj).encode("utf-8")
        result = _parse_json(content)

        assert result == [obj]


# ---------------------------------------------------------------------------
# fetch_and_parse routing for .geojson and .json URLs
# ---------------------------------------------------------------------------


class TestFetchAndParseGeoJsonRouting:
    @pytest.mark.asyncio
    async def test_routes_geojson_url_to_parse_geojson(self) -> None:
        """fetch_and_parse routes .geojson URL to _parse_geojson."""
        from mcp_canada.shared.parsers import fetch_and_parse

        fake_content = json.dumps({
            "type": "FeatureCollection",
            "features": [
                {"type": "Feature", "geometry": None, "properties": {"a": 1}},
            ],
        }).encode("utf-8")

        async def fake_cached_fetch(key: str, ttl: int, fetcher: Any) -> tuple[Any, bool]:
            data = await fetcher()
            return data, False

        mock_response = MagicMock()
        mock_response.content = fake_content
        mock_response.raise_for_status = MagicMock()

        with (
            patch("mcp_canada.shared.parsers.cached_fetch", side_effect=fake_cached_fetch),
            patch("httpx.AsyncClient") as mock_client_cls,
            patch("mcp_canada.shared.parsers._parse_geojson", return_value=[{"a": 1}]) as mock_gj,
        ):
            mock_client = AsyncMock()
            mock_client.__aenter__ = AsyncMock(return_value=mock_client)
            mock_client.__aexit__ = AsyncMock(return_value=None)
            mock_client.get = AsyncMock(return_value=mock_response)
            mock_client_cls.return_value = mock_client

            result, _ = await fetch_and_parse("https://example.com/data.geojson")

        mock_gj.assert_called_once_with(fake_content)

    @pytest.mark.asyncio
    async def test_routes_json_url_to_parse_json(self) -> None:
        """fetch_and_parse routes .json URL to _parse_json."""
        from mcp_canada.shared.parsers import fetch_and_parse

        fake_content = json.dumps([{"id": 1}]).encode("utf-8")

        async def fake_cached_fetch(key: str, ttl: int, fetcher: Any) -> tuple[Any, bool]:
            data = await fetcher()
            return data, False

        mock_response = MagicMock()
        mock_response.content = fake_content
        mock_response.raise_for_status = MagicMock()

        with (
            patch("mcp_canada.shared.parsers.cached_fetch", side_effect=fake_cached_fetch),
            patch("httpx.AsyncClient") as mock_client_cls,
            patch("mcp_canada.shared.parsers._parse_json", return_value=[{"id": 1}]) as mock_json,
        ):
            mock_client = AsyncMock()
            mock_client.__aenter__ = AsyncMock(return_value=mock_client)
            mock_client.__aexit__ = AsyncMock(return_value=None)
            mock_client.get = AsyncMock(return_value=mock_response)
            mock_client_cls.return_value = mock_client

            result, _ = await fetch_and_parse("https://example.com/data.json")

        mock_json.assert_called_once_with(fake_content)
